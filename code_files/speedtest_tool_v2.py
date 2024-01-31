from code_files.ssh_network_conf import ssh_network_conf
from copy import copy
import os

import openpyxl
from openpyxl import load_workbook
import re
import code_files.data as data

template_top_gap = 4
template_data_rows = 5
recurring_gap = 3

table_header_width = 3



from code_files.conf import *

def copy_style(source, destination):
    if source.has_style:
        destination.font = copy(source.font)
        destination.border = copy(source.border)
        destination.fill = copy(source.fill)
        destination.number_format = copy(source.number_format)
        destination.protection = copy(source.protection)
        destination.alignment = copy(source.alignment)

def copy_dims(ws1, ws2):
    for column_letter in range(65, 79):
        ws2.column_dimensions[chr(column_letter)].bestFit = True

def alpha_to_numeric(r):
    # {"start":"B5","end":"M13"}
    start_col = ord(re.search("[A-Z]+", r["start"]).group(0)) - 64
    end_col = ord(re.search("[A-Z]+", r["end"]).group(0)) - 64
    start_row = int(re.search(r"\d+", r["start"]).group(0))
    end_row = int(re.search(r"\d+", r["end"]).group(0))
    return {
        "start_row": start_row,
        "start_col": start_col,
        "end_row": end_row,
        "end_col": end_col
    }



def copy_rows(template_worksheet, output_worksheet, nrange, last_row, test_value, upload_bw, download_bw):
    for index in range(1, table_header_width + 1):
        copy_row(template_worksheet, output_worksheet,template_top_gap + index, last_row + recurring_gap + index, nrange["start_col"],
                 nrange["end_col"])
    for index in range(1, len(test_value) + 1):
        if index == len(test_value):
            copy_row(template_worksheet, output_worksheet, template_top_gap + template_data_rows + table_header_width,
                     last_row + recurring_gap + table_header_width + index,
                     nrange["start_col"], nrange["end_col"])
        else:
            copy_row(template_worksheet, output_worksheet, template_top_gap + index + table_header_width,
                     last_row + recurring_gap + table_header_width + index, nrange["start_col"], nrange["end_col"])

    merge_row = last_row + recurring_gap + 1
    merge_cells(output_worksheet, merge_row)
    add_column_data(output_worksheet, test_value, last_row + recurring_gap + table_header_width + 1)
    if len(output_worksheet["G"]) >= 15:
        upload_cell = "E" + str(last_row)
        upload_bw = float(output_worksheet[upload_cell].value) * 2
        download_cell = "E" + str(last_row)
        download_bw = float(output_worksheet[download_cell].value) * 2
    replace_bandwidth(output_worksheet, upload_bw, last_row + recurring_gap + table_header_width + 1,"up",  test_value)
    replace_bandwidth(output_worksheet, download_bw, last_row + recurring_gap + table_header_width + 1, "down", test_value)
    first_row = last_row + recurring_gap + table_header_width + 1
    update_result_set(output_worksheet, first_row, len(test_value), template_worksheet)
    for index, server in enumerate(test_value):
        output_worksheet["G" + str(index + first_row)] = "Speedtest #" + str(index + 1)



def copy_row(template_worksheet, output_worksheet, template_row, target_row, start_col, end_col):
    for col_n in range(start_col, end_col + 1):
        duplicate_cell(template_worksheet, output_worksheet, template_row, col_n, target_row, col_n)


def duplicate_cell(template_worksheet, output_worksheet, source_row, source_col, dest_row, dest_col):
    source_cell = chr(source_col + 64) + str(source_row)
    dest_cell = chr(dest_col + 64) + str(dest_row)
    output_worksheet[dest_cell] = template_worksheet[source_cell].value
    #output_worksheet[dest_cell]._style = template_worksheet[source_cell]._style
    copy_style(template_worksheet[source_cell], output_worksheet[dest_cell])


def merge_cells(worksheet, merge_row):
    merged_cells = [
        {"start_row": merge_row, "start_column": 2, "end_row": merge_row, "end_column": 4}
    ]
    for row in merged_cells:
        worksheet.merge_cells(start_row=row["start_row"], start_column=row["start_column"], end_row=row["end_row"],
                              end_column=row["end_column"])


def replace_bandwidth(worksheet, bw, row, direction, test_value):
    for i in range(0, len(test_value)):
        if direction == "up":
            column = "F"
        else:
            column = "E"
        cell = column + str(row + i)
        worksheet[cell] = bw


def add_column_data(worksheet, data, row):
    for item in data:
        worksheet["H" + str(row)] = float(item["latency"])
        worksheet["I" + str(row)] = float(item["download_speed"])
        worksheet["J" + str(row)] = float(item["upload_speed"])
        row = row + 1


def update_results(worksheet, row_int):
    row = str(row_int)
    download_bandwidth_cell = "E" + row
    upload_bandwidth_cell = "F" + row
    download_speed_cell = "I" + row
    upload_speed_cell = "J" + row
    result_cell = "M" + row
    download_bandwidth = worksheet[download_bandwidth_cell].value
    upload_bandwidth = worksheet[upload_bandwidth_cell].value
    download_speed = worksheet[download_speed_cell].value
    upload_speed = worksheet[upload_speed_cell].value
    upload_percentage = (float(upload_speed) / float(upload_bandwidth)) * 100
    download_percentage = (float(download_speed) / float(download_bandwidth)) * 100
    if upload_percentage >= 80 and download_percentage >= 80:
        worksheet[result_cell] = "Passed"
    else:
        worksheet[result_cell] = "Failed"
    download_percentage = round(download_percentage, 2)
    upload_percentage = round(upload_percentage, 2)
    worksheet["K" + row] = str(download_percentage) + "%"
    worksheet["L" + row] = str(upload_percentage) + "%"


def update_result_set(worksheet, set_start, rows_len, template_worksheet):
    for i in range(0, rows_len):
        update_results(worksheet, set_start + i)
    calculate_average(worksheet, set_start, "E", rows_len, False, template_worksheet)
    calculate_average(worksheet, set_start, "F", rows_len, False, template_worksheet)
    calculate_average(worksheet, set_start, "I", rows_len, False, template_worksheet)
    calculate_average(worksheet, set_start, "J", rows_len, False, template_worksheet)
    calculate_average(worksheet, set_start, "K", rows_len, True, template_worksheet)
    calculate_average(worksheet, set_start, "L", rows_len, True, template_worksheet)


def calculate_average(worksheet, set_start, column, rows_len, percentage, template_worksheet):
    total = 0
    for row in range(set_start, set_start + rows_len):
        cell = column + str(row)
        total = total + float(str(worksheet[cell].value).replace("%", ""))
    average = round(total / rows_len, 2)
    cell = column + str(row + 1)
    if percentage:
        worksheet[cell].value = str(average) + "%"
        if column == "K":
            kcell = "B" + str(set_start - 1)
            worksheet[kcell].value = str(average) + "%"
            copy_style(template_worksheet["K13"], worksheet[cell])
        else:
            lcell = "C" + str(set_start - 1)
            worksheet[lcell].value = str(average) + "%"
            copy_style(template_worksheet["L13"], worksheet[cell])

    else:
        worksheet[cell].value = str(average)
        copy_style(template_worksheet["E13"], worksheet[cell])


def update_result_sets(worksheet, test_value):
    for index, single_set in enumerate(test_value):
        starting_row = 8 + index * 12
        update_result_set(worksheet, starting_row)


def update_sheet(sheetname, template_file, output_file):
    template_workbook = load_workbook(template_file)
    template_worksheet = template_workbook[sheetname]
    if not os.path.isfile(output_file):
        output_workbook = openpyxl.Workbook()
        default_name = output_workbook.sheetnames[0]
        output_workbook[default_name].title = sheetname
        output_workbook[sheetname]["G3"].value = sheetname
        copy_style(template_worksheet["G3"], output_workbook[sheetname]["G3"])
        copy_dims(template_worksheet, output_workbook[sheetname])
        output_workbook.save(output_file)
    output_workbook = load_workbook(output_file)
    if sheetname in output_workbook.sheetnames:
        output_worksheet = output_workbook[sheetname]
    else:
        output_workbook.create_sheet(sheetname)
        output_worksheet = output_workbook[sheetname]
        output_workbook[sheetname]["G3"].value = sheetname
        copy_style(template_worksheet["G3"], output_workbook[sheetname]["G3"])
        copy_dims(template_worksheet, output_workbook[sheetname])
    nrange = alpha_to_numeric({"start": "B5", "end": "M13"})
    for i in range(0, test_sets):
        if mock_data == True:
            test_value = data.get_mock_data()
            if test_value:
                test_value = test_value[0:rows_per_table]
        else:
            try:
                test_value = data.get_test_value_set(rows_per_table)
            except:
                exit("Error Occurred")
        if test_value:
            copy_rows(template_worksheet, output_worksheet, nrange, len(output_worksheet["G"]), test_value, upload_bw = 5, download_bw = 10)

    output_workbook.save(output_file)


def trigger_test(network_conf):
    # output_file = "/opt/n1/v4/output.xlsx"
    output_file= "/home/user/Desktop/Speedtest/code_files/output.xlsx"
    template_path = "/home/user/Desktop/Speedtest/code_files/template.xlsx"
    update_sheet(network_conf, template_path, output_file)


######################################################################################################################
# Sohaib Changes below this point

# def execute_speedtest():
#     if not data.check_internet_connectivity():
#         print("Internet Not Reachable")
#         return False
#
#     if not mock_network_conf:
#         ssh_network_conf.check_and_configure_service("IPoE")
#     trigger_test("DHCP")
#     if not mock_network_conf:
#         ssh_network_conf.check_and_configure_service("PPPoE")
#     trigger_test("PPPoE")
#     return True

services_var = {"IPoE": {"configure_name": "IPoE", "trigger_name": "DHCP"},
                "VIPoE": {"configure_name": "PPPoE", "trigger_name": "PPPoE"}}


def check_internet_connectivity():
    if not data.check_internet_connectivity():
        print("Internet Not Reachable")
        return False


def configure_service(service):
    print("Validating configure and test for service: %s" % service)

    if not mock_network_conf:
        print("Attempting to configure service: %s" % service)
        return ssh_network_conf.check_and_configure_service(services_var[service]["configure_name"])
    else:
        print("Skip configuring services: Using mock network conf")
        return True


def speed_test(service):
    print("Attempting speedtest for service: %s" % service)
    trigger_test(services_var[service]["trigger_name"])
    return True


if __name__ == "__main__":
    check_internet_connectivity()

    configure_service("IPoE")
    speed_test("IPoE")

    configure_service("VIPoE")
    speed_test("VIPoE")

    print("All tasks complete")
